import os
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart, LineChart, Series
from openpyxl.chart.label import DataLabel, DataLabelList

from models import StockAnalysisResult
from config import config_manager


def get_timestamp():
    """Get current timestamp in YYYYMMDD_HHMMSS format"""
    return datetime.now().strftime('%Y%m%d_%H%M%S')


class OutputGenerator:
    """Generate various output reports for stock screening results"""
    
    def __init__(self):
        """Initialize the output generator"""
        self.config = config_manager.config
        self.output_settings = self.config.get('output', {})
        self.timestamp = get_timestamp()
        
    def write_text_report(self, results: List[StockAnalysisResult], total_stocks: int) -> str:
        """
        Write a text report with screening results
        
        Args:
            results: List of stock analysis results
            total_stocks: Total number of stocks analyzed
            
        Returns:
            The path to the generated file
        """
        # Create the filename
        prefix = self.output_settings.get('filename_prefix', 'nasdaq_growth_stocks')
        filename = f"{prefix}_report_{self.timestamp}.txt"
        
        with open(filename, 'w', encoding='utf-8') as f:
            # Write header
            f.write(f"NASDAQ Stock Screening Results\n")
            f.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Screened {total_stocks} stocks, found {len(results)} qualifying stocks.\n\n")
            f.write("=" * 80 + "\n\n")
            
            # Write individual stock analyses
            for i, stock in enumerate(results, 1):
                f.write(f"#{i}: {stock.symbol} - {stock.company_name}\n")
                f.write(f"Sector: {stock.sector}\n")
                f.write(f"Industry: {stock.industry}\n")
                f.write(f"Market Cap: ${stock.market_cap:,.0f}\n")
                f.write(f"Quality Score: {stock.normalized_quality_score:.4f}\n\n")
                
                # Valuation metrics
                f.write("Valuation Metrics:\n")
                f.write(f"  P/E Ratio: {stock.metrics.get('per', 0):.2f}\n")
                f.write(f"  P/B Ratio: {stock.metrics.get('pbr', 0):.2f}\n")
                f.write(f"  FCF Yield: {stock.metrics.get('fcf_yield', 0):.2%}\n\n")
                
                # Growth metrics
                f.write("Growth Metrics:\n")
                f.write(f"  Revenue CAGR: {stock.metrics.get('revenue_cagr', 0):.2%}\n")
                f.write(f"  EPS CAGR: {stock.metrics.get('eps_cagr', 0):.2%}\n")
                f.write(f"  FCF CAGR: {stock.metrics.get('fcf_cagr', 0):.2%}\n")
                f.write(f"  Latest ROE: {stock.metrics.get('latest_roe', 0):.2%}\n\n")
                
                # Risk metrics
                f.write("Risk Metrics:\n")
                f.write(f"  Debt-to-Equity: {stock.metrics.get('debt_to_equity', 0):.2f}\n")
                f.write(f"  Interest Coverage: {stock.metrics.get('interest_coverage', 0):.2f}\n\n")
                
                # Component scores
                f.write("Component Scores:\n")
                f.write(f"  Growth Quality: {stock.component_scores.get('growth_score', 0):.4f}\n")
                f.write(f"  Risk Assessment: {stock.component_scores.get('risk_score', 0):.4f}\n")
                f.write(f"  Valuation: {stock.component_scores.get('valuation_score', 0):.4f}\n")
                f.write(f"  Market Sentiment: {stock.component_scores.get('sentiment_score', 0):.4f}\n")
                f.write(f"  Coherence Multiplier: {stock.component_scores.get('coherence_multiplier', 1.0):.4f}\n\n")
                
                # Insider trading
                if stock.insider_trading and stock.insider_trading.recent_transactions:
                    f.write("Recent Insider Trading:\n")
                    f.write(f"  Buy Count: {stock.insider_trading.buy_count}\n")
                    f.write(f"  Sell Count: {stock.insider_trading.sell_count}\n")
                    f.write(f"  Buy/Sell Ratio: {stock.insider_trading.net_buy_sell_ratio:.2f}\n\n")
                
                # Earnings
                if stock.earnings_info and stock.earnings_info.latest_eps_actual is not None:
                    f.write("Latest Earnings:\n")
                    f.write(f"  EPS Actual: {stock.earnings_info.latest_eps_actual:.2f}\n")
                    f.write(f"  EPS Estimated: {stock.earnings_info.latest_eps_estimated:.2f}\n")
                    if stock.earnings_info.eps_surprise_percentage is not None:
                        f.write(f"  EPS Surprise: {stock.earnings_info.eps_surprise_percentage:.2%}\n")
                    if stock.earnings_info.next_earnings_date:
                        f.write(f"  Next Earnings Date: {stock.earnings_info.next_earnings_date}\n\n")
                
                # Sector percentiles
                if hasattr(stock, 'sector_percentile') and stock.sector_percentile:
                    f.write("Sector Percentiles:\n")
                    for metric, percentile in stock.sector_percentile.items():
                        metric_name = metric.split('.')[-1]
                        f.write(f"  {metric_name}: {percentile:.0f}th percentile\n")
                    f.write("\n")
                
                # Separator between stocks
                f.write("-" * 80 + "\n\n")
        
        logging.info(f"Text report written to {filename}")
        return filename
    
    def write_excel_report(self, results: List[StockAnalysisResult], total_stocks: int) -> str:
        """
        Write an Excel report with screening results
        
        Args:
            results: List of stock analysis results
            total_stocks: Total number of stocks analyzed
            
        Returns:
            The path to the generated file
        """
        # Create the filename
        prefix = self.output_settings.get('filename_prefix', 'nasdaq_growth_stocks')
        filename = f"{prefix}_report_{self.timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Create summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        self._write_summary_sheet(summary_sheet, results, total_stocks)
        
        # Create detail sheet
        detail_sheet = wb.create_sheet("Details")
        self._write_detail_sheet(detail_sheet, results)
        
        # Create growth analysis sheet
        growth_sheet = wb.create_sheet("Growth Analysis")
        self._write_growth_sheet(growth_sheet, results)
        
        # Create risk analysis sheet
        risk_sheet = wb.create_sheet("Risk Analysis")
        self._write_risk_sheet(risk_sheet, results)
        
        # Create valuation sheet
        valuation_sheet = wb.create_sheet("Valuation")
        self._write_valuation_sheet(valuation_sheet, results)
        
        # Create sector analysis sheet
        sector_sheet = wb.create_sheet("Sector Analysis")
        self._write_sector_sheet(sector_sheet, results)
        
        # Save the workbook
        wb.save(filename)
        
        logging.info(f"Excel report written to {filename}")
        return filename
    
    def _write_summary_sheet(self, sheet, results: List[StockAnalysisResult], total_stocks: int):
        """Write the summary sheet with key metrics"""
        # Set column widths
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Header row style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Add title
        sheet.merge_cells('A1:I1')
        sheet['A1'] = f"NASDAQ Stock Screening Results - {datetime.now().strftime('%Y-%m-%d')}"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = Alignment(horizontal="center")
        
        # Add summary stats
        sheet.merge_cells('A3:D3')
        sheet['A3'] = f"Total Stocks Screened: {total_stocks}"
        sheet['A3'].font = Font(bold=True)
        
        sheet.merge_cells('A4:D4')
        sheet['A4'] = f"Qualifying Stocks: {len(results)}"
        sheet['A4'].font = Font(bold=True)
        
        # Headers
        headers = ["Rank", "Symbol", "Company", "Sector", "Market Cap", "P/E", "ROE", "Growth Score", "Quality Score"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Data rows
        for row, stock in enumerate(results, 7):
            rank = row - 6
            sheet.cell(row=row, column=1, value=rank)
            
            # Add hyperlink to symbol for Yahoo Finance
            symbol_cell = sheet.cell(row=row, column=2, value=stock.symbol)
            symbol_cell.hyperlink = f"https://finance.yahoo.com/quote/{stock.symbol}"
            symbol_cell.font = Font(color="0000FF", underline="single")
            
            sheet.cell(row=row, column=3, value=stock.company_name)
            sheet.cell(row=row, column=4, value=stock.sector)
            sheet.cell(row=row, column=5, value=stock.market_cap)
            sheet.cell(row=row, column=5).number_format = "$#,##0"
            sheet.cell(row=row, column=6, value=stock.metrics.get('per', 0))
            sheet.cell(row=row, column=6).number_format = "0.00"
            sheet.cell(row=row, column=7, value=stock.metrics.get('latest_roe', 0))
            sheet.cell(row=row, column=7).number_format = "0.00%"
            sheet.cell(row=row, column=8, value=stock.component_scores.get('growth_score', 0))
            sheet.cell(row=row, column=8).number_format = "0.00"
            sheet.cell(row=row, column=9, value=stock.normalized_quality_score)
            sheet.cell(row=row, column=9).number_format = "0.00"
            
            # Conditional formatting for quality score
            if stock.normalized_quality_score >= 0.8:
                sheet.cell(row=row, column=9).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif stock.normalized_quality_score >= 0.6:
                sheet.cell(row=row, column=9).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Freeze header row
        sheet.freeze_panes = 'A7'
        
        # Add auto-filter to data range
        if len(results) > 0:
            last_row = 6 + len(results)
            sheet.auto_filter.ref = f"A6:I{last_row}"
        
        # Create charts
        self._add_sector_distribution_chart(sheet, results)
    
    def _write_detail_sheet(self, sheet, results: List[StockAnalysisResult]):
        """Write the detailed metrics sheet"""
        # Set column widths
        for col in range(1, 30):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Header row style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Symbol", "Company", "Sector", "Industry", "Market Cap",
            "Quality Score", "Growth Score", "Risk Score", "Valuation Score", "Sentiment Score",
            "Revenue CAGR", "EPS CAGR", "FCF CAGR", "Latest ROE", "Avg ROE",
            "P/E Ratio", "P/B Ratio", "FCF Yield", "Debt/Equity", "Interest Coverage",
            "Coherence Multiplier", "Insider Buy/Sell", "EPS Surprise"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Data rows
        for row, stock in enumerate(results, 2):
            col = 1
            sheet.cell(row=row, column=col, value=stock.symbol); col += 1
            sheet.cell(row=row, column=col, value=stock.company_name); col += 1
            sheet.cell(row=row, column=col, value=stock.sector); col += 1
            sheet.cell(row=row, column=col, value=stock.industry); col += 1
            
            sheet.cell(row=row, column=col, value=stock.market_cap)
            sheet.cell(row=row, column=col).number_format = "$#,##0"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.normalized_quality_score)
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.component_scores.get('growth_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.component_scores.get('risk_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.component_scores.get('valuation_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.component_scores.get('sentiment_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('revenue_cagr', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('eps_cagr', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('fcf_cagr', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('latest_roe', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('avg_roe', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('per', 0))
            sheet.cell(row=row, column=col).number_format = "0.0"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('pbr', 0))
            sheet.cell(row=row, column=col).number_format = "0.0"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('fcf_yield', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('debt_to_equity', 0))
            sheet.cell(row=row, column=col).number_format = "0.00"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('interest_coverage', 0))
            sheet.cell(row=row, column=col).number_format = "0.00"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.component_scores.get('coherence_multiplier', 1.0))
            sheet.cell(row=row, column=col).number_format = "0.00"; col += 1
            
            # Insider trading
            buy_sell_ratio = 0
            if stock.insider_trading:
                buy_sell_ratio = stock.insider_trading.net_buy_sell_ratio
            sheet.cell(row=row, column=col, value=buy_sell_ratio)
            sheet.cell(row=row, column=col).number_format = "0.00"; col += 1
            
            # Earnings surprise
            eps_surprise = 0
            if stock.earnings_info and stock.earnings_info.eps_surprise_percentage is not None:
                eps_surprise = stock.earnings_info.eps_surprise_percentage
            sheet.cell(row=row, column=col, value=eps_surprise)
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
        
        # Add auto-filter to all columns
        if len(results) > 0:
            last_row = 1 + len(results)
            last_col_letter = get_column_letter(len(headers))
            sheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
        
        # Apply conditional formatting to score columns
        if len(results) > 0:
            # Quality Score column (column 6)
            for row in range(2, 2 + len(results)):
                quality_score = sheet.cell(row=row, column=6).value
                if quality_score and quality_score >= 0.8:
                    sheet.cell(row=row, column=6).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif quality_score and quality_score >= 0.6:
                    sheet.cell(row=row, column=6).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif quality_score and quality_score < 0.4:
                    sheet.cell(row=row, column=6).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def _write_growth_sheet(self, sheet, results: List[StockAnalysisResult]):
        """Write the growth analysis sheet"""
        # Set column widths
        for col in range(1, 15):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Header row style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Symbol", "Company", "Sector", "Growth Score",
            "Revenue CAGR", "EPS CAGR", "FCF CAGR",
            "Revenue Consistency", "EPS Consistency", "FCF Consistency",
            "Sustainability Score", "Magnitude Score", "Consistency Score"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Data rows
        for row, stock in enumerate(results, 2):
            growth_analysis = stock.growth_analysis
            
            col = 1
            sheet.cell(row=row, column=col, value=stock.symbol); col += 1
            sheet.cell(row=row, column=col, value=stock.company_name); col += 1
            sheet.cell(row=row, column=col, value=stock.sector); col += 1
            
            sheet.cell(row=row, column=col, value=stock.component_scores.get('growth_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=growth_analysis.get('revenue_cagr', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            sheet.cell(row=row, column=col, value=growth_analysis.get('eps_cagr', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            sheet.cell(row=row, column=col, value=growth_analysis.get('fcf_cagr', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            if 'consistency_scores' in growth_analysis:
                consistency_scores = growth_analysis['consistency_scores']
                
                sheet.cell(row=row, column=col, value=consistency_scores.get('revenue', 0))
                sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
                
                sheet.cell(row=row, column=col, value=consistency_scores.get('eps', 0))
                sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
                
                sheet.cell(row=row, column=col, value=consistency_scores.get('fcf', 0))
                sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            else:
                col += 3  # Skip consistency columns
            
            sheet.cell(row=row, column=col, value=growth_analysis.get('sustainability_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=growth_analysis.get('magnitude_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=growth_analysis.get('consistency_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
        
        # Create growth comparison chart
        self._add_growth_comparison_chart(sheet, results)
    
    def _write_risk_sheet(self, sheet, results: List[StockAnalysisResult]):
        """Write the risk analysis sheet"""
        # Set column widths
        for col in range(1, 15):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Header row style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Symbol", "Company", "Sector", "Risk Score",
            "Debt Score", "Working Capital Score", "Margin Stability", "Cash Flow Quality",
            "Debt/Equity", "Interest Coverage", "Debt/EBITDA"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Data rows
        for row, stock in enumerate(results, 2):
            risk_assessment = stock.risk_assessment
            
            col = 1
            sheet.cell(row=row, column=col, value=stock.symbol); col += 1
            sheet.cell(row=row, column=col, value=stock.company_name); col += 1
            sheet.cell(row=row, column=col, value=stock.sector); col += 1
            
            sheet.cell(row=row, column=col, value=stock.component_scores.get('risk_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=risk_assessment.get('debt_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=risk_assessment.get('working_capital_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=risk_assessment.get('margin_stability_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=risk_assessment.get('cash_flow_quality_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('debt_to_equity', 0))
            sheet.cell(row=row, column=col).number_format = "0.00"; col += 1
            
            sheet.cell(row=row, column=col, value=stock.metrics.get('interest_coverage', 0))
            sheet.cell(row=row, column=col).number_format = "0.00"; col += 1
            
            debt_to_ebitda = 0
            if hasattr(stock, 'metrics') and 'debt_to_ebitda' in stock.metrics:
                debt_to_ebitda = stock.metrics['debt_to_ebitda']
            sheet.cell(row=row, column=col, value=debt_to_ebitda)
            sheet.cell(row=row, column=col).number_format = "0.00"; col += 1
    
    def _write_valuation_sheet(self, sheet, results: List[StockAnalysisResult]):
        """Write the valuation analysis sheet"""
        # Set column widths
        for col in range(1, 15):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Header row style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Symbol", "Company", "Sector", "Valuation Score",
            "P/E Ratio", "P/B Ratio", "FCF Yield", 
            "P/E Score", "P/B Score", "FCF Yield Score", "Growth-Adjusted Score"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Data rows
        for row, stock in enumerate(results, 2):
            valuation_analysis = stock.valuation_analysis
            
            col = 1
            sheet.cell(row=row, column=col, value=stock.symbol); col += 1
            sheet.cell(row=row, column=col, value=stock.company_name); col += 1
            sheet.cell(row=row, column=col, value=stock.sector); col += 1
            
            sheet.cell(row=row, column=col, value=stock.component_scores.get('valuation_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=valuation_analysis.get('per', 0))
            sheet.cell(row=row, column=col).number_format = "0.0"; col += 1
            
            sheet.cell(row=row, column=col, value=valuation_analysis.get('pbr', 0))
            sheet.cell(row=row, column=col).number_format = "0.0"; col += 1
            
            sheet.cell(row=row, column=col, value=valuation_analysis.get('fcf_yield', 0))
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            sheet.cell(row=row, column=col, value=valuation_analysis.get('per_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=valuation_analysis.get('pbr_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=valuation_analysis.get('fcf_yield_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=valuation_analysis.get('growth_adjusted_score', 0))
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
        
        # Create valuation comparison chart
        self._add_valuation_comparison_chart(sheet, results)
    
    def _write_sector_sheet(self, sheet, results: List[StockAnalysisResult]):
        """Write the sector analysis sheet"""
        # Set column widths
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Group stocks by sector
        sector_groups = {}
        for stock in results:
            sector = stock.sector
            if sector not in sector_groups:
                sector_groups[sector] = []
            sector_groups[sector].append(stock)
        
        # Calculate sector average metrics
        sector_metrics = {}
        for sector, stocks in sector_groups.items():
            if not stocks:
                continue
                
            # Calculate averages
            avg_quality = sum(stock.normalized_quality_score for stock in stocks) / len(stocks)
            avg_growth = sum(stock.component_scores.get('growth_score', 0) for stock in stocks) / len(stocks)
            avg_risk = sum(stock.component_scores.get('risk_score', 0) for stock in stocks) / len(stocks)
            avg_valuation = sum(stock.component_scores.get('valuation_score', 0) for stock in stocks) / len(stocks)
            
            avg_pe = sum(stock.metrics.get('per', 0) for stock in stocks) / len(stocks)
            avg_roe = sum(stock.metrics.get('latest_roe', 0) for stock in stocks) / len(stocks)
            
            sector_metrics[sector] = {
                'count': len(stocks),
                'avg_quality': avg_quality,
                'avg_growth': avg_growth,
                'avg_risk': avg_risk,
                'avg_valuation': avg_valuation,
                'avg_pe': avg_pe,
                'avg_roe': avg_roe
            }
        
        # Header row style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Sector", "Count", "Avg Quality", "Avg Growth", "Avg Risk", 
            "Avg Valuation", "Avg P/E", "Avg ROE"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Data rows
        row = 2
        for sector, metrics in sector_metrics.items():
            col = 1
            sheet.cell(row=row, column=col, value=sector); col += 1
            sheet.cell(row=row, column=col, value=metrics['count']); col += 1
            
            sheet.cell(row=row, column=col, value=metrics['avg_quality'])
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=metrics['avg_growth'])
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=metrics['avg_risk'])
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=metrics['avg_valuation'])
            sheet.cell(row=row, column=col).number_format = "0.000"; col += 1
            
            sheet.cell(row=row, column=col, value=metrics['avg_pe'])
            sheet.cell(row=row, column=col).number_format = "0.0"; col += 1
            
            sheet.cell(row=row, column=col, value=metrics['avg_roe'])
            sheet.cell(row=row, column=col).number_format = "0.0%"; col += 1
            
            row += 1
        
        # Create sector comparison charts
        self._add_sector_quality_chart(sheet, sector_metrics)
    
    def _add_sector_distribution_chart(self, sheet, results: List[StockAnalysisResult]):
        """Add a sector distribution pie chart to the sheet"""
        # Group stocks by sector
        sector_counts = {}
        for stock in results:
            sector = stock.sector
            if sector not in sector_counts:
                sector_counts[sector] = 0
            sector_counts[sector] += 1
        
        # Write sector counts for chart
        sheet.cell(row=3, column=11, value="Sector")
        sheet.cell(row=3, column=12, value="Count")
        
        for i, (sector, count) in enumerate(sector_counts.items(), 4):
            sheet.cell(row=i, column=11, value=sector)
            sheet.cell(row=i, column=12, value=count)
        
        # Create pie chart
        pie = PieChart()
        labels = Reference(sheet, min_col=11, min_row=4, max_row=3 + len(sector_counts))
        data = Reference(sheet, min_col=12, min_row=3, max_row=3 + len(sector_counts))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Sector Distribution"
        
        # Show data labels as percentage
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        
        # Add the chart to the sheet
        sheet.add_chart(pie, "A20")
    
    def _add_growth_comparison_chart(self, sheet, results: List[StockAnalysisResult]):
        """Add a growth comparison chart to the sheet"""
        # Get top stocks by growth score
        top_stocks = sorted(results, key=lambda x: x.component_scores.get('growth_score', 0), reverse=True)[:10]
        
        # Prepare data for chart
        row_offset = len(results) + 5
        sheet.cell(row=row_offset, column=1, value="Symbol")
        sheet.cell(row=row_offset, column=2, value="Revenue CAGR")
        sheet.cell(row=row_offset, column=3, value="EPS CAGR")
        sheet.cell(row=row_offset, column=4, value="FCF CAGR")
        
        for i, stock in enumerate(top_stocks, 1):
            sheet.cell(row=row_offset + i, column=1, value=stock.symbol)
            sheet.cell(row=row_offset + i, column=2, value=stock.growth_analysis.get('revenue_cagr', 0))
            sheet.cell(row=row_offset + i, column=3, value=stock.growth_analysis.get('eps_cagr', 0))
            sheet.cell(row=row_offset + i, column=4, value=stock.growth_analysis.get('fcf_cagr', 0))
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Growth Comparison - Top 10 Growth Stocks"
        chart.y_axis.title = "CAGR"
        chart.x_axis.title = "Stock"
        
        data = Reference(sheet, min_col=2, max_col=4, min_row=row_offset, max_row=row_offset + len(top_stocks))
        cats = Reference(sheet, min_col=1, min_row=row_offset + 1, max_row=row_offset + len(top_stocks))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add the chart to the sheet
        sheet.add_chart(chart, "A" + str(row_offset + 15))
    
    def _add_valuation_comparison_chart(self, sheet, results: List[StockAnalysisResult]):
        """Add a valuation comparison chart to the sheet"""
        # Get top stocks by valuation score
        top_stocks = sorted(results, key=lambda x: x.component_scores.get('valuation_score', 0), reverse=True)[:10]
        
        # Prepare data for chart
        row_offset = len(results) + 5
        sheet.cell(row=row_offset, column=1, value="Symbol")
        sheet.cell(row=row_offset, column=2, value="P/E Ratio")
        sheet.cell(row=row_offset, column=3, value="FCF Yield")
        
        for i, stock in enumerate(top_stocks, 1):
            sheet.cell(row=row_offset + i, column=1, value=stock.symbol)
            sheet.cell(row=row_offset + i, column=2, value=stock.valuation_analysis.get('per', 0))
            sheet.cell(row=row_offset + i, column=3, value=stock.valuation_analysis.get('fcf_yield', 0))
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Valuation Comparison - Top 10 Value Stocks"
        chart.y_axis.title = "Ratio"
        chart.x_axis.title = "Stock"
        
        data = Reference(sheet, min_col=2, max_col=2, min_row=row_offset, max_row=row_offset + len(top_stocks))
        cats = Reference(sheet, min_col=1, min_row=row_offset + 1, max_row=row_offset + len(top_stocks))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add the chart to the sheet
        sheet.add_chart(chart, "A" + str(row_offset + 15))
        
        # Create FCF yield chart
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "FCF Yield Comparison - Top 10 Value Stocks"
        chart2.y_axis.title = "FCF Yield"
        chart2.x_axis.title = "Stock"
        
        data2 = Reference(sheet, min_col=3, max_col=3, min_row=row_offset, max_row=row_offset + len(top_stocks))
        cats2 = Reference(sheet, min_col=1, min_row=row_offset + 1, max_row=row_offset + len(top_stocks))
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        
        # Add the second chart to the sheet
        sheet.add_chart(chart2, "H" + str(row_offset + 15))
    
    def _add_sector_quality_chart(self, sheet, sector_metrics):
        """Add a sector quality comparison chart to the sheet"""
        # Prepare data for chart
        row_offset = len(sector_metrics) + 5
        sheet.cell(row=row_offset, column=1, value="Sector")
        sheet.cell(row=row_offset, column=2, value="Quality Score")
        sheet.cell(row=row_offset, column=3, value="Growth Score")
        sheet.cell(row=row_offset, column=4, value="Risk Score")
        sheet.cell(row=row_offset, column=5, value="Valuation Score")
        
        for i, (sector, metrics) in enumerate(sector_metrics.items(), 1):
            sheet.cell(row=row_offset + i, column=1, value=sector)
            sheet.cell(row=row_offset + i, column=2, value=metrics['avg_quality'])
            sheet.cell(row=row_offset + i, column=3, value=metrics['avg_growth'])
            sheet.cell(row=row_offset + i, column=4, value=metrics['avg_risk'])
            sheet.cell(row=row_offset + i, column=5, value=metrics['avg_valuation'])
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Sector Quality Comparison"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Sector"
        
        data = Reference(sheet, min_col=2, max_col=5, min_row=row_offset, max_row=row_offset + len(sector_metrics))
        cats = Reference(sheet, min_col=1, min_row=row_offset + 1, max_row=row_offset + len(sector_metrics))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add the chart to the sheet
        sheet.add_chart(chart, "A" + str(row_offset + 15))
    
    def write_csv_report(self, results: List[StockAnalysisResult], total_stocks: int) -> str:
        """
        Write a CSV report with screening results
        
        Args:
            results: List of stock analysis results
            total_stocks: Total number of stocks analyzed
            
        Returns:
            The path to the generated file
        """
        import csv
        
        # Create the filename
        prefix = self.output_settings.get('filename_prefix', 'nasdaq_growth_stocks')
        filename = f"{prefix}_report_{self.timestamp}.csv"
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            # Define CSV headers
            fieldnames = [
                'Rank', 'Symbol', 'Company Name', 'Sector', 'Industry', 'Market Cap',
                'Quality Score', 'Growth Score', 'Risk Score', 'Valuation Score', 'Sentiment Score',
                'P/E Ratio', 'P/B Ratio', 'FCF Yield', 'Revenue CAGR', 'EPS CAGR', 'FCF CAGR',
                'Latest ROE', 'Debt-to-Equity', 'Interest Coverage', 'Insider Buy Count',
                'Insider Sell Count', 'Buy/Sell Ratio', 'Latest EPS', 'EPS Surprise %'
            ]
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            # Write data rows
            for i, stock in enumerate(results, 1):
                row = {
                    'Rank': i,
                    'Symbol': stock.symbol,
                    'Company Name': stock.company_name,
                    'Sector': stock.sector,
                    'Industry': stock.industry,
                    'Market Cap': stock.market_cap,
                    'Quality Score': round(stock.normalized_quality_score, 4),
                    'Growth Score': round(stock.component_scores.get('growth_score', 0), 4),
                    'Risk Score': round(stock.component_scores.get('risk_score', 0), 4),
                    'Valuation Score': round(stock.component_scores.get('valuation_score', 0), 4),
                    'Sentiment Score': round(stock.component_scores.get('sentiment_score', 0), 4),
                    'P/E Ratio': round(stock.metrics.get('per', 0), 2),
                    'P/B Ratio': round(stock.metrics.get('pbr', 0), 2),
                    'FCF Yield': round(stock.metrics.get('fcf_yield', 0) * 100, 2),
                    'Revenue CAGR': round(stock.metrics.get('revenue_cagr', 0) * 100, 2),
                    'EPS CAGR': round(stock.metrics.get('eps_cagr', 0) * 100, 2),
                    'FCF CAGR': round(stock.metrics.get('fcf_cagr', 0) * 100, 2),
                    'Latest ROE': round(stock.metrics.get('latest_roe', 0) * 100, 2),
                    'Debt-to-Equity': round(stock.metrics.get('debt_to_equity', 0), 2),
                    'Interest Coverage': round(stock.metrics.get('interest_coverage', 0), 2)
                }
                
                # Add insider trading info if available
                if stock.insider_trading:
                    row['Insider Buy Count'] = stock.insider_trading.buy_count
                    row['Insider Sell Count'] = stock.insider_trading.sell_count
                    row['Buy/Sell Ratio'] = round(stock.insider_trading.net_buy_sell_ratio, 2)
                else:
                    row['Insider Buy Count'] = 0
                    row['Insider Sell Count'] = 0
                    row['Buy/Sell Ratio'] = 0
                
                # Add earnings info if available
                if stock.earnings_info and stock.earnings_info.latest_eps_actual is not None:
                    row['Latest EPS'] = round(stock.earnings_info.latest_eps_actual, 2)
                    if stock.earnings_info.eps_surprise_percentage is not None:
                        row['EPS Surprise %'] = round(stock.earnings_info.eps_surprise_percentage * 100, 2)
                    else:
                        row['EPS Surprise %'] = 0
                else:
                    row['Latest EPS'] = 0
                    row['EPS Surprise %'] = 0
                
                writer.writerow(row)
        
        logging.info(f"CSV report written to {filename}")
        return filename
    
    def write_json_report(self, results: List[StockAnalysisResult], total_stocks: int) -> str:
        """
        Write a JSON report with screening results
        
        Args:
            results: List of stock analysis results
            total_stocks: Total number of stocks analyzed
            
        Returns:
            The path to the generated file
        """
        import json
        
        # Create the filename
        prefix = self.output_settings.get('filename_prefix', 'nasdaq_growth_stocks')
        filename = f"{prefix}_report_{self.timestamp}.json"
        
        # Prepare data for JSON export
        data = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'total_stocks_screened': total_stocks,
                'qualifying_stocks': len(results),
                'config': {
                    'initial_filters': self.config.get('initial_filters', {}),
                    'roe_criteria': self.config.get('roe_criteria', {}),
                    'growth_targets': self.config.get('growth_targets', {}),
                    'scoring_weights': self.config.get('scoring_weights', {})
                }
            },
            'results': []
        }
        
        # Add stock data
        for i, stock in enumerate(results, 1):
            stock_data = {
                'rank': i,
                'symbol': stock.symbol,
                'company_name': stock.company_name,
                'sector': stock.sector,
                'industry': stock.industry,
                'market_cap': stock.market_cap,
                'scores': {
                    'quality_score': stock.normalized_quality_score,
                    'growth_score': stock.component_scores.get('growth_score', 0),
                    'risk_score': stock.component_scores.get('risk_score', 0),
                    'valuation_score': stock.component_scores.get('valuation_score', 0),
                    'sentiment_score': stock.component_scores.get('sentiment_score', 0),
                    'coherence_multiplier': stock.component_scores.get('coherence_multiplier', 1.0)
                },
                'metrics': stock.metrics,
                'financial_metrics': stock.financial_metrics.__dict__ if stock.financial_metrics else {},
                'insider_trading': {
                    'buy_count': stock.insider_trading.buy_count if stock.insider_trading else 0,
                    'sell_count': stock.insider_trading.sell_count if stock.insider_trading else 0,
                    'net_buy_sell_ratio': stock.insider_trading.net_buy_sell_ratio if stock.insider_trading else 0
                },
                'earnings': {
                    'latest_eps_actual': stock.earnings_info.latest_eps_actual if stock.earnings_info else None,
                    'latest_eps_estimated': stock.earnings_info.latest_eps_estimated if stock.earnings_info else None,
                    'eps_surprise_percentage': stock.earnings_info.eps_surprise_percentage if stock.earnings_info else None,
                    'next_earnings_date': stock.earnings_info.next_earnings_date if stock.earnings_info else None
                }
            }
            data['results'].append(stock_data)
        
        # Write JSON file
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)
        
        logging.info(f"JSON report written to {filename}")
        return filename


# Create the global output generator instance
output_generator = OutputGenerator()